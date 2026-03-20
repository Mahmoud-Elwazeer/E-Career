"""
View to download a sample CSV/XLSX template for job import.
"""
import io
from datetime import date
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAdminUser


SAMPLE_ROWS = [
    {
        "title": "Senior Python Developer",
        "company_name": "TechCorp Egypt",
        "location": "Cairo, Egypt",
        "location_type": "hybrid",
        "industry": "technology",
        "experience_level": "senior",
        "description": "We are looking for an experienced Python developer to join our backend team. You will work with Django, PostgreSQL, and Redis to build scalable APIs.",
        "source_url": "https://linkedin.com/jobs/example-1",
        "posted_at": str(date.today()),
        "status": "active",
        "salary_min": "15000",
        "salary_max": "25000",
        "salary_currency": "EGP",
        "source_name": "",
        "tags": "Python, Django, PostgreSQL, REST API",
        "deadline": "",
        "slug": "",
    },
    {
        "title": "Frontend React Developer",
        "company_name": "Digital Agency",
        "location": "Remote",
        "location_type": "remote",
        "industry": "technology",
        "experience_level": "mid",
        "description": "Join our frontend team to build modern web applications using React, TypeScript, and Tailwind CSS.",
        "source_url": "https://wuzzuf.net/jobs/example-2",
        "posted_at": str(date.today()),
        "status": "active",
        "salary_min": "10000",
        "salary_max": "18000",
        "salary_currency": "EGP",
        "source_name": "",
        "tags": "React, TypeScript, Tailwind CSS, JavaScript",
        "deadline": "",
        "slug": "",
    },
]

HEADERS = [
    "title", "company_name", "location", "location_type", "industry",
    "experience_level", "description", "source_url", "posted_at", "status",
    "salary_min", "salary_max", "salary_currency", "source_name", "tags",
    "deadline", "slug",
]


class JobTemplateDownloadView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        fmt = request.query_params.get("format", "csv").lower()

        if fmt == "xlsx":
            return self._xlsx_response()
        return self._csv_response()

    def _csv_response(self):
        import csv
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=HEADERS)
        writer.writeheader()
        for row in SAMPLE_ROWS:
            writer.writerow(row)
        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="jobs_import_template.csv"'
        return response

    def _xlsx_response(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()

        # ── Sheet 1: Template ──
        ws = wb.active
        ws.title = "Jobs"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Headers
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Sample data
        for row_idx, row_data in enumerate(SAMPLE_ROWS, 2):
            for col_idx, header in enumerate(HEADERS, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                cell.border = thin_border

        # Auto-width
        for col_idx, header in enumerate(HEADERS, 1):
            max_len = max(len(header), max(len(str(r.get(header, ""))) for r in SAMPLE_ROWS))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

        # ── Sheet 2: Guide ──
        guide = wb.create_sheet("Guide")
        guide_data = [
            ["Column", "Required?", "Allowed Values", "Notes"],
            ["title", "YES", "Any text", "Job title"],
            ["company_name", "YES", "Any text", "Auto-created if new"],
            ["location", "YES", "Any text", 'e.g. "Cairo, Egypt" or "Remote"'],
            ["location_type", "YES", "remote / onsite / hybrid", ""],
            ["industry", "YES", "technology / finance / healthcare / education / marketing / engineering / design / sales / other", ""],
            ["experience_level", "YES", "entry / mid / senior / lead", ""],
            ["description", "YES", "Any text", "Full job description"],
            ["source_url", "YES", "URL", "Link to original job posting"],
            ["posted_at", "YES", "YYYY-MM-DD", "Date the job was posted"],
            ["status", "no", "active / pending / archived", "Default: active"],
            ["salary_min", "no", "Number", ""],
            ["salary_max", "no", "Number", ""],
            ["salary_currency", "no", "e.g. USD, EGP", "Default: USD"],
            ["source_name", "no", "Existing source name", "Must already exist in Sources"],
            ["tags", "no", "Comma-separated", 'e.g. "Python, Django, REST API"'],
            ["deadline", "no", "YYYY-MM-DD", "Application deadline"],
            ["slug", "no", "URL slug", "Auto-generated from title if blank"],
        ]

        guide_header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        for row_idx, row in enumerate(guide_data, 1):
            for col_idx, val in enumerate(row, 1):
                cell = guide.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if row_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = guide_header_fill
                elif col_idx == 2 and val == "YES":
                    cell.font = Font(bold=True, color="DC2626")

        for col_idx in range(1, 5):
            guide.column_dimensions[guide.cell(row=1, column=col_idx).column_letter].width = [20, 12, 50, 40][col_idx - 1]

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="jobs_import_template.xlsx"'
        return response
